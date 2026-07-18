"""
src/analytics/valuation.py
===========================
Sprint 4 Day 26 — Valuation Module.

Computes:
  - FCF Yield = free_cash_flow_cr / market_cap_crore * 100
  - Sector median P/E for each broad_sector (latest year)
  - Overvaluation flags:
      Caution  — P/E > sector_median_PE * 1.5
      Discount — P/E < sector_median_PE * 0.7
      Fair     — otherwise
  - 5-year median P/E per company from market_cap table

Outputs:
  - output/valuation_summary.xlsx
  - output/valuation_flags.csv

Run:
    python -m src.analytics.valuation
    python src/analytics/valuation.py
"""

from __future__ import annotations

import os
import sqlite3
import sys
from pathlib import Path

import pandas as pd

# ---------------------------------------------------------------------------
# Path setup — works when run from repo root or as module
# ---------------------------------------------------------------------------
_REPO_ROOT = Path(__file__).resolve().parents[2]
DB_PATH    = str(_REPO_ROOT / "data" / "nifty100.db")
OUTPUT_DIR = _REPO_ROOT / "output"


def _connect() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def load_valuation_data() -> pd.DataFrame:
    """
    Load combined valuation dataset — latest financial_ratios + 2024 market_cap.
    """
    sql = """
        WITH latest_fr AS (
            SELECT company_id, MAX(year) AS latest_year
            FROM financial_ratios
            GROUP BY company_id
        )
        SELECT
            fr.company_id,
            TRIM(c.company_name)  AS company_name,
            s.broad_sector,
            mc.pe_ratio,
            mc.pb_ratio,
            mc.ev_ebitda          AS ev_ebitda,
            mc.dividend_yield_pct,
            mc.market_cap_crore,
            fr.free_cash_flow_cr,
            fr.composite_quality_score
        FROM latest_fr lf
        JOIN financial_ratios fr
            ON fr.company_id = lf.company_id AND fr.year = lf.latest_year
        JOIN companies c ON c.id = fr.company_id
        LEFT JOIN sectors s ON s.company_id = fr.company_id
        LEFT JOIN market_cap mc
            ON mc.company_id = fr.company_id AND mc.year = 2024
        ORDER BY s.broad_sector, fr.composite_quality_score DESC
    """
    with _connect() as conn:
        return pd.read_sql_query(sql, conn)


def load_5yr_median_pe() -> pd.DataFrame:
    """
    Compute 5-year median P/E per company from market_cap table (years 2019-2024).
    """
    sql = """
        SELECT company_id, pe_ratio
        FROM market_cap
        WHERE year BETWEEN 2019 AND 2024
          AND pe_ratio IS NOT NULL
          AND pe_ratio > 0
    """
    with _connect() as conn:
        df = pd.read_sql_query(sql, conn)
    return df.groupby("company_id")["pe_ratio"].median().reset_index().rename(
        columns={"pe_ratio": "pe_5yr_median"}
    )


# ---------------------------------------------------------------------------
# Computation
# ---------------------------------------------------------------------------

def compute_fcf_yield(df: pd.DataFrame) -> pd.DataFrame:
    """FCF Yield (%) = FCF / Market Cap * 100"""
    df = df.copy()
    df["fcf_yield_pct"] = (
        df["free_cash_flow_cr"] / df["market_cap_crore"].replace(0, float("nan"))
    ) * 100
    return df


def compute_sector_median_pe(df: pd.DataFrame) -> pd.DataFrame:
    """Compute sector median P/E and merge back into df."""
    sector_pe = (
        df.dropna(subset=["pe_ratio", "broad_sector"])
        .groupby("broad_sector")["pe_ratio"]
        .median()
        .reset_index()
        .rename(columns={"pe_ratio": "sector_median_pe"})
    )
    return df.merge(sector_pe, on="broad_sector", how="left")


def apply_valuation_flags(df: pd.DataFrame) -> pd.DataFrame:
    """
    Flag each company vs its sector median P/E:
        Caution  — P/E > sector_median_PE * 1.5
        Discount — P/E < sector_median_PE * 0.7
        Fair     — otherwise (including NaN P/E)
    """
    df = df.copy()

    def _flag(row: pd.Series) -> str:
        pe   = row.get("pe_ratio")
        smed = row.get("sector_median_pe")
        if pd.isna(pe) or pd.isna(smed) or smed == 0:
            return "Fair"
        if pe > smed * 1.5:
            return "Caution"
        if pe < smed * 0.7:
            return "Discount"
        return "Fair"

    df["flag"] = df.apply(_flag, axis=1)
    return df


def compute_pe_vs_sector_pct(df: pd.DataFrame) -> pd.DataFrame:
    """PE_vs_sector_median_pct = (PE / sector_median_PE - 1) * 100"""
    df = df.copy()
    df["pe_vs_sector_median_pct"] = (
        (df["pe_ratio"] / df["sector_median_pe"].replace(0, float("nan")) - 1) * 100
    ).round(1)
    return df


# ---------------------------------------------------------------------------
# Output generation
# ---------------------------------------------------------------------------

SUMMARY_COLS = [
    "company_id", "company_name", "broad_sector",
    "pe_ratio", "pb_ratio", "ev_ebitda", "fcf_yield_pct",
    "pe_5yr_median", "pe_vs_sector_median_pct", "sector_median_pe", "flag",
    "composite_quality_score",
]

FLAGS_COLS = [
    "company_id", "company_name", "broad_sector",
    "pe_ratio", "sector_median_pe", "pe_vs_sector_median_pct",
    "flag", "market_cap_crore", "pb_ratio", "fcf_yield_pct",
    "composite_quality_score",
]


def write_outputs(summary: pd.DataFrame) -> None:
    """Write valuation_summary.xlsx and valuation_flags.csv to output/."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # ── valuation_summary.xlsx ──────────────────────────────────────────────
    xlsx_path = OUTPUT_DIR / "valuation_summary.xlsx"
    export_cols = [c for c in SUMMARY_COLS if c in summary.columns]
    out_df = summary[export_cols].copy()

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            out_df.to_excel(writer, index=False, sheet_name="Valuation Summary")
            ws = writer.sheets["Valuation Summary"]

            # Header style
            hdr_fill = PatternFill("solid", fgColor="1A1F2E")
            hdr_font = Font(bold=True, color="4F8EF7")
            for cell in ws[1]:
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal="center")

            # Flag colour coding
            flag_col_idx = export_cols.index("flag") + 1
            fill_caution  = PatternFill("solid", fgColor="FFCC00")
            fill_discount = PatternFill("solid", fgColor="22C55E")
            fill_fair     = PatternFill("solid", fgColor="FFFFFF")

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                flag_cell = row[flag_col_idx - 1]
                if flag_cell.value == "Caution":
                    flag_cell.fill = fill_caution
                    flag_cell.font = Font(bold=True, color="7C2D12")
                elif flag_cell.value == "Discount":
                    flag_cell.fill = fill_discount
                    flag_cell.font = Font(bold=True, color="064E3B")

            # Auto-width
            for col_cells in ws.columns:
                length = max(len(str(c.value or "")) for c in col_cells) + 4
                ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(length, 30)

        print(f"OK: valuation_summary.xlsx written ({len(out_df)} rows)")

    except ImportError:
        out_df.to_excel(xlsx_path, index=False)
        print(f"OK: valuation_summary.xlsx written (basic, openpyxl styling unavailable)")

    # ── valuation_flags.csv ─────────────────────────────────────────────────
    flagged    = summary[summary["flag"].isin(["Caution", "Discount"])].copy()
    flags_cols = [c for c in FLAGS_COLS if c in flagged.columns]
    csv_path   = OUTPUT_DIR / "valuation_flags.csv"
    flagged[flags_cols].to_csv(csv_path, index=False)
    print(f"OK: valuation_flags.csv written ({len(flagged)} flagged companies)")


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def run() -> pd.DataFrame:
    """Run the full valuation pipeline and return the summary DataFrame."""
    print("Loading valuation data…")
    df = load_valuation_data()
    print(f"  Loaded {len(df)} companies")

    print("Computing FCF yield…")
    df = compute_fcf_yield(df)

    print("Computing sector median P/E…")
    df = compute_sector_median_pe(df)

    print("Applying valuation flags…")
    df = apply_valuation_flags(df)

    print("Computing P/E vs sector median %…")
    df = compute_pe_vs_sector_pct(df)

    print("Loading 5-year median P/E…")
    pe5 = load_5yr_median_pe()
    df  = df.merge(pe5, on="company_id", how="left")
    df  = df.rename(columns={"pe_5yr_median": "pe_5yr_median"})

    print("Writing output files...")
    write_outputs(df)

    # Summary stats
    flag_counts = df["flag"].value_counts()
    print("\n-- Flag Summary --")
    for flag, cnt in flag_counts.items():
        print(f"  {flag:10s}: {cnt} companies")
    print("------------------")

    return df


if __name__ == "__main__":
    result = run()
    print(f"\nDone - {len(result)} companies processed.")
