"""
src/screener/exporter.py
========================
Sprint 3 Day 17 — Generate output/screener_output.xlsx.

One sheet per preset, sorted by composite_quality_score descending.
Cells are colour-coded green (meets threshold) / red (fails threshold).

Requires: openpyxl
"""

from __future__ import annotations

import os
from pathlib import Path
from typing import Any

import pandas as pd

try:
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False

from src.screener.presets import PRESETS, PresetScreener
from src.screener.scorer import add_scores_to_df

# Colours
GREEN_FILL = "C6EFCE"   # Excel green
RED_FILL   = "FFC7CE"   # Excel red
HEADER_FILL = "1F4E79"  # Dark blue header
HEADER_FONT_COLOUR = "FFFFFF"

# Columns to export per sheet
EXPORT_COLUMNS = [
    "company_name",
    "broad_sector",
    "return_on_equity_pct",
    "roce_pct",
    "net_profit_margin_pct",
    "debt_to_equity",
    "interest_coverage",
    "free_cash_flow_cr",
    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "eps_cagr_5yr",
    "operating_profit_margin_pct",
    "asset_turnover",
    "cfo_quality_score",
    "capital_allocation_pattern",
    "icr_label",
    "composite_quality_score",
    "sector_relative_score",
    "dividend_yield_pct",
    "dividend_payout_ratio_pct",
]

COLUMN_LABELS = {
    "company_name":               "Company",
    "broad_sector":               "Sector",
    "return_on_equity_pct":       "ROE %",
    "roce_pct":                   "ROCE %",
    "net_profit_margin_pct":      "NPM %",
    "debt_to_equity":             "D/E",
    "interest_coverage":          "ICR",
    "free_cash_flow_cr":          "FCF (Cr)",
    "revenue_cagr_5yr":           "Rev CAGR 5yr %",
    "pat_cagr_5yr":               "PAT CAGR 5yr %",
    "eps_cagr_5yr":               "EPS CAGR 5yr %",
    "operating_profit_margin_pct":"OPM %",
    "asset_turnover":             "Asset Turnover",
    "cfo_quality_score":          "CFO Quality",
    "capital_allocation_pattern": "Cap Alloc",
    "icr_label":                  "ICR Label",
    "composite_quality_score":    "Quality Score",
    "sector_relative_score":      "Sector Score",
    "dividend_yield_pct":         "Div Yield %",
    "dividend_payout_ratio_pct":  "Div Payout %",
}


# ---------------------------------------------------------------------------
# Threshold checker — does a cell value meet the preset threshold?
# ---------------------------------------------------------------------------

def _meets_threshold(
    col: str,
    value: Any,
    preset_name: str,
) -> bool | None:
    """
    Return True if value meets the preset threshold for this column.
    Return None if this column has no threshold in the preset.
    """
    preset = PRESETS.get(preset_name, {})
    filters = preset.get("filters", {})

    # Map column name back to filter key
    col_to_filter = {
        "return_on_equity_pct":       "roe_min",
        "debt_to_equity":             "de_max",
        "free_cash_flow_cr":          "fcf_min",
        "revenue_cagr_5yr":           "revenue_cagr_5yr_min",
        "pat_cagr_5yr":               "pat_cagr_5yr_min",
        "operating_profit_margin_pct":"opm_min",
        "pe_ratio":                   "pe_max",
        "pb_ratio":                   "pb_max",
        "dividend_yield_pct":         "dividend_yield_min",
        "interest_coverage":          "icr_min",
        "net_profit_margin_pct":      "net_profit_min",
        "eps_cagr_5yr":               "eps_cagr_5yr_min",
        "asset_turnover":             "asset_turnover_min",
        "sales":                      "sales_min",
    }

    filter_key = col_to_filter.get(col)
    if filter_key not in filters:
        return None

    threshold = filters[filter_key]
    if threshold is None or value is None:
        return None

    try:
        v = float(value)
    except (TypeError, ValueError):
        return None

    # Determine direction from filter key suffix
    if filter_key.endswith("_min"):
        return v >= threshold
    else:
        return v <= threshold


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

def generate_screener_excel(
    preset_results: dict[str, pd.DataFrame],
    output_path: str = "output/screener_output.xlsx",
) -> str:
    """
    Write one sheet per preset to screener_output.xlsx.

    Parameters
    ----------
    preset_results : dict of {preset_name: filtered DataFrame}
    output_path    : destination path

    Returns path written to.
    """
    if not _OPENPYXL:
        raise ImportError("openpyxl is required: pip install openpyxl")

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)

    green = PatternFill(start_color=GREEN_FILL, end_color=GREEN_FILL, fill_type="solid")
    red   = PatternFill(start_color=RED_FILL,   end_color=RED_FILL,   fill_type="solid")
    hdr_fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
    hdr_font = Font(color=HEADER_FONT_COLOUR, bold=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for preset_name, df in preset_results.items():
            if df.empty:
                # Write empty sheet with headers
                empty = pd.DataFrame(columns=EXPORT_COLUMNS)
                empty.to_excel(writer, sheet_name=_sheet_name(preset_name), index=False)
                continue

            # Select and reorder columns (skip missing)
            cols = [c for c in EXPORT_COLUMNS if c in df.columns]
            export_df = df[cols].copy()

            # Round numeric columns
            for col in export_df.select_dtypes(include="number").columns:
                export_df[col] = export_df[col].round(2)

            sheet_name = _sheet_name(preset_name)
            export_df.to_excel(writer, sheet_name=sheet_name, index=False)

            ws = writer.sheets[sheet_name]

            # Style header row
            for cell in ws[1]:
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal="center")
                # Rename header to friendly label
                if cell.value in COLUMN_LABELS:
                    cell.value = COLUMN_LABELS[cell.value]

            # Colour-code data rows
            for row_idx, (_, row) in enumerate(export_df.iterrows(), start=2):
                for col_idx, col in enumerate(cols, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    val = row.get(col)
                    meets = _meets_threshold(col, val, preset_name)
                    if meets is True:
                        cell.fill = green
                    elif meets is False:
                        cell.fill = red

            # Auto-width columns
            for col_idx, col in enumerate(cols, start=1):
                col_letter = get_column_letter(col_idx)
                max_len = max(
                    len(COLUMN_LABELS.get(col, col)),
                    export_df[col].astype(str).str.len().max() if not export_df.empty else 0,
                )
                ws.column_dimensions[col_letter].width = min(max_len + 2, 30)

    return output_path


def _sheet_name(preset_name: str) -> str:
    """Convert preset_name to a valid Excel sheet name (max 31 chars)."""
    label = PRESETS.get(preset_name, {}).get("label", preset_name)
    return label[:31]


# ---------------------------------------------------------------------------
# Convenience runner — load from DB and generate Excel
# ---------------------------------------------------------------------------

def run_screener_export(
    db_path: str,
    output_path: str = "output/screener_output.xlsx",
) -> str:
    """
    Full pipeline: load DB → score → run all presets → export Excel.

    Returns path to generated file.
    """
    ps = PresetScreener()
    ps.load_from_db(db_path)

    # Add composite scores to the full universe DataFrame
    scored_df = add_scores_to_df(ps._engine._df)
    ps._engine._df = scored_df
    ps._df = scored_df

    preset_results = ps.run_all()
    return generate_screener_excel(preset_results, output_path)
