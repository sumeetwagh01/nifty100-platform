"""
src/analytics/peer_comparison.py
=================================
Sprint 3 Day 20 — Generate output/peer_comparison.xlsx.

11 sheets — one per peer group.
Each sheet:
  - Columns: company_id, company_name + 20 metric columns + percentile rank per metric
  - Colour-coded percentile rank cells:
      green  >= 75th percentile
      yellow  25th–75th percentile
      red    <= 25th percentile
  - Benchmark company row highlighted gold/amber
  - Summary row at bottom showing peer group median per metric

Requires: openpyxl, pandas
"""

from __future__ import annotations

import os
import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Colour constants
# ---------------------------------------------------------------------------

GREEN_FILL  = PatternFill("solid", fgColor="C6EFCE")   # >= 75th
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")   # 25th–75th
RED_FILL    = PatternFill("solid", fgColor="FFC7CE")   # <= 25th
GOLD_FILL   = PatternFill("solid", fgColor="FFD700")   # benchmark row
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")   # dark blue header
MEDIAN_FILL = PatternFill("solid", fgColor="D9E1F2")   # light blue median row

HEADER_FONT  = Font(color="FFFFFF", bold=True, size=10)
BOLD_FONT    = Font(bold=True, size=10)
MEDIAN_FONT  = Font(bold=True, italic=True, size=10)
DEFAULT_FONT = Font(size=10)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

# ---------------------------------------------------------------------------
# Metric definitions — 10 metrics × (value col + rank col) = 20 columns
# ---------------------------------------------------------------------------

METRICS = [
    {"label": "ROE %",           "value_col": "return_on_equity_pct",  "rank_col": "ROE"},
    {"label": "ROCE %",          "value_col": "roce_pct",              "rank_col": "ROCE"},
    {"label": "NPM %",           "value_col": "net_profit_margin_pct", "rank_col": "Net Profit Margin"},
    {"label": "D/E",             "value_col": "debt_to_equity",        "rank_col": "D/E"},
    {"label": "FCF (Cr)",        "value_col": "free_cash_flow_cr",     "rank_col": "FCF"},
    {"label": "PAT CAGR 5yr %",  "value_col": "pat_cagr_5yr",         "rank_col": "PAT CAGR 5yr"},
    {"label": "Rev CAGR 5yr %",  "value_col": "revenue_cagr_5yr",     "rank_col": "Revenue CAGR 5yr"},
    {"label": "EPS CAGR 5yr %",  "value_col": "eps_cagr_5yr",         "rank_col": "EPS CAGR 5yr"},
    {"label": "ICR",             "value_col": "interest_coverage",     "rank_col": "Interest Coverage"},
    {"label": "Asset Turnover",  "value_col": "asset_turnover",        "rank_col": "Asset Turnover"},
]


# ---------------------------------------------------------------------------
# Data loader
# ---------------------------------------------------------------------------

def load_peer_group_data(
    conn: sqlite3.Connection,
    peer_group_name: str,
    year: int | None = None,
) -> pd.DataFrame:
    """
    Load financial_ratios + peer_percentiles for a peer group.

    Returns wide DataFrame: one row per company with value cols + rank cols.
    """
    year_filter = f"AND pp.year = {year}" if year else ""

    # Get all companies and their metric values
    value_query = f"""
        SELECT
            fr.company_id,
            c.company_name,
            fr.return_on_equity_pct,
            fr.roce_pct,
            fr.net_profit_margin_pct,
            fr.debt_to_equity,
            fr.free_cash_flow_cr,
            fr.pat_cagr_5yr,
            fr.revenue_cagr_5yr,
            fr.eps_cagr_5yr,
            fr.interest_coverage,
            fr.asset_turnover,
            fr.composite_quality_score
        FROM financial_ratios fr
        JOIN companies c ON c.id = fr.company_id
        WHERE fr.company_id IN (
            SELECT DISTINCT company_id FROM peer_percentiles
            WHERE peer_group_name = ? {year_filter}
        )
        AND fr.year = (
            SELECT MAX(year) FROM financial_ratios WHERE company_id = fr.company_id
        )
        ORDER BY fr.company_id
    """
    values_df = pd.read_sql_query(value_query, conn, params=(peer_group_name,))

    if values_df.empty:
        return pd.DataFrame()

    # Get percentile ranks — pivot to wide
    rank_query = f"""
        SELECT company_id, metric, percentile_rank
        FROM peer_percentiles
        WHERE peer_group_name = ? {year_filter}
    """
    ranks_df = pd.read_sql_query(rank_query, conn, params=(peer_group_name,))
    ranks_wide = ranks_df.pivot_table(
        index="company_id", columns="metric", values="percentile_rank"
    ).reset_index()
    ranks_wide.columns.name = None

    # Rename rank columns to avoid clash with value columns
    rank_rename = {m["rank_col"]: f"rank_{m['rank_col']}" for m in METRICS}
    ranks_wide = ranks_wide.rename(columns=rank_rename)

    merged = values_df.merge(ranks_wide, on="company_id", how="left")
    return merged


def get_peer_groups(conn: sqlite3.Connection) -> list[str]:
    """Return all distinct peer group names from peer_percentiles."""
    rows = conn.execute(
        "SELECT DISTINCT peer_group_name FROM peer_percentiles ORDER BY peer_group_name"
    ).fetchall()
    return [r[0] for r in rows]


def get_benchmark_company(
    conn: sqlite3.Connection,
    peer_group_name: str,
) -> int | None:
    """
    Identify benchmark company for a peer group — highest composite_quality_score.
    Returns company_id or None.
    """
    row = conn.execute("""
        SELECT fr.company_id
        FROM financial_ratios fr
        WHERE fr.company_id IN (
            SELECT DISTINCT company_id FROM peer_percentiles
            WHERE peer_group_name = ?
        )
        ORDER BY fr.composite_quality_score DESC
        LIMIT 1
    """, (peer_group_name,)).fetchone()
    return row[0] if row else None


# ---------------------------------------------------------------------------
# Sheet writer
# ---------------------------------------------------------------------------

def _write_peer_sheet(
    ws,
    df: pd.DataFrame,
    peer_group_name: str,
    benchmark_company_id: int | None,
) -> None:
    """Write one peer group sheet with all formatting."""

    if df.empty:
        ws.cell(1, 1).value = "No data available"
        return

    # Build column list: company_id, company_name, then pairs of (value, rank)
    base_cols = ["company_id", "company_name"]
    metric_col_pairs = []
    for m in METRICS:
        val_col  = m["value_col"]
        rank_col = f"rank_{m['rank_col']}"
        metric_col_pairs.append((m["label"], val_col, rank_col))

    # --- Header row 1: group labels spanning value+rank pairs ---
    ws.row_dimensions[1].height = 20
    ws.cell(1, 1).value = "Company ID"
    ws.cell(1, 2).value = "Company Name"
    for h_col, (label, _, _) in enumerate(metric_col_pairs, start=3):
        # Merge two cells for each metric label
        val_col_idx  = 3 + (h_col - 3) * 2
        rank_col_idx = val_col_idx + 1
        ws.merge_cells(
            start_row=1, start_column=val_col_idx,
            end_row=1,   end_column=rank_col_idx,
        )
        cell = ws.cell(1, val_col_idx)
        cell.value     = label
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER

    # Style first two header cells
    for c in [1, 2]:
        ws.merge_cells(
            start_row=1, start_column=c, end_row=2, end_column=c
        )
        cell = ws.cell(1, c)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER

    # --- Header row 2: Value / Rank sub-headers ---
    ws.row_dimensions[2].height = 16
    for h_col, (_, _, _) in enumerate(metric_col_pairs):
        val_col_idx  = 3 + h_col * 2
        rank_col_idx = val_col_idx + 1
        for ci, sub in [(val_col_idx, "Value"), (rank_col_idx, "Rank %")]:
            cell = ws.cell(2, ci)
            cell.value     = sub
            cell.fill      = HEADER_FILL
            cell.font      = HEADER_FONT
            cell.alignment = CENTER

    # --- Data rows ---
    df_sorted = df.sort_values("composite_quality_score", ascending=False).reset_index(drop=True)

    for row_idx, (_, row) in enumerate(df_sorted.iterrows(), start=3):
        is_benchmark = (row.get("company_id") == benchmark_company_id)
        ws.row_dimensions[row_idx].height = 16

        # company_id
        c1 = ws.cell(row_idx, 1, int(row["company_id"]))
        c1.alignment = CENTER
        c1.font      = BOLD_FONT if is_benchmark else DEFAULT_FONT

        # company_name
        c2 = ws.cell(row_idx, 2, str(row["company_name"]))
        c2.alignment = LEFT
        c2.font      = BOLD_FONT if is_benchmark else DEFAULT_FONT

        if is_benchmark:
            c1.fill = GOLD_FILL
            c2.fill = GOLD_FILL

        # metric pairs
        for h_col, (_, val_col, rank_col) in enumerate(metric_col_pairs):
            val_col_idx  = 3 + h_col * 2
            rank_col_idx = val_col_idx + 1

            # Value cell
            raw_val = row.get(val_col)
            val = round(float(raw_val), 2) if pd.notna(raw_val) else ""
            vc = ws.cell(row_idx, val_col_idx, val)
            vc.alignment = CENTER
            vc.font      = DEFAULT_FONT
            if is_benchmark:
                vc.fill = GOLD_FILL

            # Rank cell — colour coded
            raw_rank = row.get(rank_col)
            rank_val = round(float(raw_rank), 1) if pd.notna(raw_rank) else ""
            rc = ws.cell(row_idx, rank_col_idx, rank_val)
            rc.alignment = CENTER
            rc.font      = DEFAULT_FONT

            if isinstance(rank_val, float):
                if rank_val >= 75:
                    rc.fill = GREEN_FILL
                elif rank_val <= 25:
                    rc.fill = RED_FILL
                else:
                    rc.fill = YELLOW_FILL
            elif is_benchmark:
                rc.fill = GOLD_FILL

    # --- Median summary row ---
    median_row = row_idx + 2
    ws.row_dimensions[median_row].height = 18

    ws.cell(median_row, 1).value = "MEDIAN"
    ws.cell(median_row, 2).value = peer_group_name

    for cell in [ws.cell(median_row, 1), ws.cell(median_row, 2)]:
        cell.fill      = MEDIAN_FILL
        cell.font      = MEDIAN_FONT
        cell.alignment = CENTER

    for h_col, (_, val_col, rank_col) in enumerate(metric_col_pairs):
        val_col_idx  = 3 + h_col * 2
        rank_col_idx = val_col_idx + 1

        num_vals = pd.to_numeric(df_sorted.get(val_col,  pd.Series()), errors="coerce")
        num_rnks = pd.to_numeric(df_sorted.get(rank_col, pd.Series()), errors="coerce")

        med_val  = round(num_vals.median(), 2) if not num_vals.dropna().empty else ""
        med_rank = round(num_rnks.median(), 1) if not num_rnks.dropna().empty else ""

        for ci, val in [(val_col_idx, med_val), (rank_col_idx, med_rank)]:
            cell = ws.cell(median_row, ci, val)
            cell.fill      = MEDIAN_FILL
            cell.font      = MEDIAN_FONT
            cell.alignment = CENTER

    # --- Column widths ---
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 24
    for h_col in range(len(metric_col_pairs)):
        val_col_idx  = 3 + h_col * 2
        rank_col_idx = val_col_idx + 1
        ws.column_dimensions[get_column_letter(val_col_idx)].width  = 12
        ws.column_dimensions[get_column_letter(rank_col_idx)].width = 10

    # Freeze panes — keep headers and company name visible
    ws.freeze_panes = "C3"


# ---------------------------------------------------------------------------
# Main generator
# ---------------------------------------------------------------------------

def generate_peer_comparison_excel(
    conn: sqlite3.Connection,
    output_path: str = "output/peer_comparison.xlsx",
    year: int | None = None,
) -> str:
    """
    Generate output/peer_comparison.xlsx with 11 sheets — one per peer group.

    Returns path written to.
    """
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)

    peer_groups = get_peer_groups(conn)
    if not peer_groups:
        raise ValueError("No peer groups found in peer_percentiles table. Run PeerEngine first.")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for group_name in peer_groups:
            df = load_peer_group_data(conn, group_name, year=year)
            benchmark_id = get_benchmark_company(conn, group_name)

            # Sheet name max 31 chars
            sheet_name = group_name[:31]
            # Write empty shell so openpyxl creates the sheet
            pd.DataFrame().to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]

            _write_peer_sheet(ws, df, group_name, benchmark_id)

    return output_path
